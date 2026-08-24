"""
Загрузка супервайзеров и их агентов из файла SUPERVISORS.xlsx в Postgres.

Что делает:
  1. читает файл (колонка = бренд, внутри супервайзеры и их агенты);
  2. определяет способ записи для каждой колонки — в файле их три;
  3. проставляет агентам колонку supervisor;
  4. заводит супервайзерам учётные записи (логин = код, role='supervisor');
  5. пишет отчёт: кто остался без супервайзера, кто без агентов, дубли.

ПЕРЕД ЗАПУСКОМ выполнить supervisor_setup.sql.

ЗАПУСК:
    pip install openpyxl psycopg2-binary
    python migrate_supervisors.py "postgresql://user:pass@host:port/railway"

Файл SUPERVISORS.xlsx должен лежать рядом со скриптом.
"""

import csv
import re
import sys
from collections import defaultdict, Counter

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

SRC = "SUPERVISORS.xlsx"
REPORT = "otchet_supervayzerov.csv"

# Пароль по умолчанию для новых учёток супервайзеров.
# Каждый меняет его сам при первом входе в панель.
DEFAULT_PASSWORD = "123"


is_sv = lambda x: bool(re.match(r'^[A-Z]{2}[A-Z]*S[A-Z]*\d+$', x)) and 'S' in x[2:]
num    = lambda x: int(re.search(r'(\d+)$', x).group(1)) if re.search(r'(\d+)$', x) else 0
series = lambda x: (re.search(r'(\d+)$', x).group(1)[:2] if re.search(r'(\d+)$', x) else None)

def positional(items, direction):
    res, bucket, cur = defaultdict(list), [], None
    if direction == 'after':
        for x in items:
            if is_sv(x):
                res[x].extend(bucket); bucket = []
            else: bucket.append(x)
    else:
        for x in items:
            if is_sv(x): cur = x
            elif cur: res[cur].append(x)
            else: bucket.append(x)
    return res, bucket

def score(res):
    hits = total = 0
    for sv, ags in res.items():
        for a in ags:
            total += 1
            if series(a) and series(a) == series(sv): hits += 1
    return hits / total if total else 0

def by_series(items):
    """Агенты распределяются по серии; внутри серии — поровну между супервайзерами."""
    svs = [x for x in items if is_sv(x)]
    ags = [x for x in items if not is_sv(x)]
    res, left = defaultdict(list), []
    groups = defaultdict(list)
    for a in ags: groups[series(a)].append(a)
    sv_by_series = defaultdict(list)
    for s in svs: sv_by_series[series(s)].append(s)

    for s, alist in groups.items():
        owners = sorted(sv_by_series.get(s, []), key=num)
        if not owners:
            left.extend(alist); continue
        alist = sorted(alist, key=num)
        size = -(-len(alist) // len(owners))          # округление вверх
        for i, owner in enumerate(owners):
            res[owner].extend(alist[i * size:(i + 1) * size])
    for s in svs:
        res.setdefault(s, [])
    return res, left


def read_columns(path):
    ws = openpyxl.load_workbook(path, data_only=True).active
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    cols = {}
    for c in range(1, ws.max_column + 1):
        if not hdr[c - 1]:
            continue
        vals = [ws.cell(r, c).value for r in range(2, ws.max_row + 1)]
        cols[str(hdr[c - 1]).strip()] = [str(v).strip().upper() for v in vals if v and str(v).strip()]
    return cols


def build_mapping(cols):
    mapping, sv_brand, issues, summary = {}, {}, [], []

    for col, items in cols.items():
        svs = [x for x in items if is_sv(x)]
        pos = [i for i, x in enumerate(items) if is_sv(x)]
        # Если все супервайзеры собраны в конце колонки, порядок ничего
        # не говорит о принадлежности — распределяем по серии кода
        tail = bool(pos) and min(pos) > len(items) * 0.6

        best, method, left = None, None, []
        if not tail:
            for d in ("after", "before"):
                r, l = positional(items, d)
                if best is None or score(r) > score(best):
                    best, method, left = r, d, l

        if tail or (best is not None and score(best) < 0.5):
            best, left = by_series(items)
            method = "по серии кода"
        else:
            method = "супервайзер " + ("после агентов" if method == "after" else "перед агентами")

        for sv, ags in best.items():
            sv_brand[sv] = sv[:2]
            for a in ags:
                if a in mapping and mapping[a] != sv:
                    issues.append((col, a, f"встречается у двух супервайзеров: {mapping[a]} и {sv}"))
                mapping[a] = sv

        for a in left:
            issues.append((col, a, "нет супервайзера в файле"))
        empty = sorted([sv for sv, ags in best.items() if not ags])
        for sv in empty:
            issues.append((col, sv, "супервайзер без агентов"))

        summary.append((col, len(items) - len(svs), len(svs), method, len(left)))

    return mapping, sv_brand, issues, summary


def main():
    if len(sys.argv) < 2:
        print('Использование: python migrate_supervisors.py "<строка подключения>"')
        sys.exit(1)

    try:
        cols = read_columns(SRC)
    except FileNotFoundError:
        print(f"❌ Файл {SRC} не найден — положите его рядом со скриптом.")
        sys.exit(1)

    mapping, sv_brand, issues, summary = build_mapping(cols)

    print(f"{'колонка':10} {'агентов':>8} {'суперв.':>8}  {'способ разбора':30} {'без sv':>7}")
    for row in summary:
        print(f"{row[0]:10} {row[1]:>8} {row[2]:>8}  {row[3]:30} {row[4]:>7}")

    print(f"\nСвязок агент → супервайзер: {len(mapping)}")
    print(f"Супервайзеров: {len(sv_brand)}")

    print("\nПодключаюсь к Postgres...")
    try:
        conn = psycopg2.connect(sys.argv[1], connect_timeout=15)
    except psycopg2.OperationalError as e:
        print(f"❌ Не удалось подключиться: {e}")
        print("   Строку подключения бери в Railway → Postgres → Variables → DATABASE_PUBLIC_URL")
        sys.exit(1)

    cur = conn.cursor()
    cur.execute("""
        SELECT column_name FROM information_schema.columns
         WHERE table_name = 'users' AND column_name = 'supervisor'
    """)
    if cur.fetchone() is None:
        print("\n❌ В таблице users нет колонки supervisor.")
        print("   Сначала выполни supervisor_setup.sql.")
        conn.close()
        sys.exit(1)

    # ---- учётки супервайзеров ----
    sv_rows = [(sv.lower(), DEFAULT_PASSWORD, "supervisor", brand, None)
               for sv, brand in sorted(sv_brand.items())]
    execute_values(cur, """
        INSERT INTO users (login, password, role, brand, supervisor)
        VALUES %s
        ON CONFLICT (login) DO UPDATE
            SET role  = 'supervisor',
                brand = EXCLUDED.brand
    """, sv_rows)

    # ---- агенты: заводим отсутствующих и проставляем супервайзера ----
    ag_rows = [(a.lower(), DEFAULT_PASSWORD, "agent", a[:2], sv)
               for a, sv in sorted(mapping.items())]
    execute_values(cur, """
        INSERT INTO users (login, password, role, brand, supervisor)
        VALUES %s
        ON CONFLICT (login) DO UPDATE
            SET supervisor = EXCLUDED.supervisor,
                brand      = EXCLUDED.brand
    """, ag_rows)
    conn.commit()

    # ---- что в базе, но не в файле ----
    cur.execute("""
        SELECT upper(login) FROM users
         WHERE COALESCE(role, 'agent') = 'agent' AND supervisor IS NULL
         ORDER BY login
    """)
    orphans = [r[0] for r in cur.fetchall()]

    cur.execute("SELECT count(*) FROM users WHERE role = 'supervisor'")
    total_sv = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM users WHERE supervisor IS NOT NULL")
    total_ag = cur.fetchone()[0]
    cur.close()
    conn.close()

    for a in orphans:
        issues.append(("база", a, "агент есть в базе, но не найден в файле"))

    with open(REPORT, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Источник", "Код", "Замечание"])
        w.writerows(issues)

    print(f"\n✅ Супервайзеров в базе: {total_sv}")
    print(f"✅ Агентов с супервайзером: {total_ag}")
    print(f"⚠️  Замечаний: {len(issues)}")
    for reason, count in Counter(i[2] for i in issues).most_common():
        print(f"     • {reason}: {count}")
    print(f"\n📄 Подробности: {REPORT}")
    print(f"🔑 Пароль новых учёток: {DEFAULT_PASSWORD} — смените после первого входа")


if __name__ == "__main__":
    main()
